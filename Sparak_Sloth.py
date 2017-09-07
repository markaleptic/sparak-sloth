import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter.filedialog import askopenfilename
from tkinter.simpledialog import askinteger
from PIL import Image, ImageTk
import pyautogui
from colour import Color
import openpyxl
from openpyxl.cell import get_column_letter
import os
import datetime
import time
import timeit

###################################
#           CONSTANTS             #
###################################
WRK_DIR = 'C:/Program Files/Sparak Accounting Sloth'
LARGE_FONT = ("Verdana", 12)
NORM_FONT = ("Verdana", 10)
SMALL_FONT = ("Verdana", 8)
BACKGROUND_COLOR = Color("#FCEBAE")
TIME_INTERVAL = 0.002
# Sparak Tran Codes
DEBIT_TRAN_CODE = [1, 7, 9, 12, 19, 23, 24, 29, 31, 35, 39, 41, 45, 47, 49, 56, 59, 62, 64, 67, 69, 73]
CREDIT_TRAN_CODE = [2, 4, 6, 8, 11, 18, 22, 28, 30, 36, 40, 42, 46, 48, 50, 55, 58, 61, 63, 66, 68, 72]

###################################
#        Global Variables         #
###################################
debit_entry_total = 0.00
credit_entry_total = 0.00
input_file_name = "No File Selected"
input_transaction_count = 0

###################################################
# Method receives a string message that user needs
# to see and shows via popup message box. Popup 
# disappears after user clicks okay.
###################################################
def popupmsg(msg):
    popup = tk.Tk()
    popup.wm_title("Sloth Message")
    label = ttk.Label(popup, text=msg, font=NORM_FONT)
    label.pack(side="top", fill="x", pady=10)
    B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
    B1.pack()
    popup.lift()
    popup.mainloop()

class Sparak_Sloth(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        tk.Tk.iconbitmap(self, default="slothicon.ico")
        tk.Tk.wm_title(self, "Slothful Accounting")

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (StartPage, EntryPage): # All pages have to go in the frames list for frames to pull up window
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")
        self.show_frame(StartPage)

    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()

class StartPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
# --------------------- Background Image ---------------------
        width, height = 1024, 640
        image = Image.open('slothBackground.png')
        if image.size != (width, height):
            image = image.resize((width, height), Image.ANTIALIAS)

        image = ImageTk.PhotoImage(image)
        bg_label = tk.Label(self, image = image)
        bg_label.place(x=0, y=0, relwidth=1, relheight=1)
        bg_label.image = image
# --------------------- Background Image ---------------------

        # Header
        label = tk.Label(self, text=("Sparak Accounting Sloth"), bg=BACKGROUND_COLOR, font=LARGE_FONT)
        label.pack(pady=10,padx=10) 
        # Buttons
        button1 = ttk.Button(self, text="Enter Transactions", command=lambda: controller.show_frame(EntryPage))
        button1.pack(pady=1, ipadx=6)
        button2 = ttk.Button(self, text="Quit Application", command=sys.exit)
        button2.pack(pady=1, ipadx=12)   
    
class EntryPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
# --------------------- Background Image ---------------------
        width, height = 1024, 640
        image = Image.open('slothBackground.png')
        if image.size != (width, height):
            image = image.resize((width, height), Image.ANTIALIAS)

        image = ImageTk.PhotoImage(image)
        bg_label = tk.Label(self, image = image)
        bg_label.place(x=0, y=0, relwidth=1, relheight=1)
        bg_label.image = image
# ------------------------------------------------------------
        #########################################################
        # GUI method creates box to see input file, entries, and 
        # credit / debit totals
        #########################################################
        def set_transaction_text(self):
            label_text = (  "Input File Name: \t" + input_file_name + 
                          "\nEntry Count:\t" + str(int(input_transaction_count)) + 
                          "\nDebit Total:\t$" + format(debit_entry_total, ',.2f') + 
                          "\nCredit Total:\t$" + format(credit_entry_total, ',.2f'))
            return label_text
        
        # Frame header text
        header_label = tk.Label(self, text=("Enter Sparak Transactions"), bg=BACKGROUND_COLOR, font=LARGE_FONT)
        header_label.grid(row=0, column=0, sticky=NSEW, columnspan=10)
                
        # Label to entry data: input file name, number of entries to make, and the debit / credit 
        entry_data_label = tk.Label(self, text=set_transaction_text(self),justify=LEFT, relief=GROOVE)
        entry_data_label.grid(row=1, column=1, rowspan=4, columnspan=1, sticky=NE, pady=6, ipadx=5)
        
        # Set weights for grid to format buttons
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)      
        self.grid_columnconfigure(2, weight=1)
        self.grid_rowconfigure(2, weight=1)  
        self.grid_columnconfigure(3, weight=1)
        self.grid_rowconfigure(3, weight=1)
        self.grid_columnconfigure(4, weight=1)
        self.grid_rowconfigure(4, weight=1)
        # Buttons
        button1 = ttk.Button(self, text="Load Transactions", command=lambda: selectFile(self))
        button1.grid(row=1, column=2, ipadx=10, padx=5, sticky=E)
        button2 = ttk.Button(self, text="Enter Transactions", command=lambda: enter_into_sparak(self, paymentArray))
        button2.grid(row=2, column=2, ipadx=9, padx=5, sticky=E) 
        button3 = ttk.Button(self, text="Clear Loaded Transactions", command=lambda: clear_payment_box(self))
        button3.grid(row=1, column=3, ipadx=1,sticky=W)
        button4 = ttk.Button(self, text="Delete Sparak Transactions", command=lambda: delete_sparak_entries(self))
        button4.grid(row=2, column=3, sticky=W)
        button5 = ttk.Button(self, text="Return to Main Menu", command=lambda: controller.show_frame(StartPage))
        button5.grid(row=3, column=3, ipadx=14, sticky=W)
 
        # Tree Formatting
        tree = ttk.Treeview(self)
        tree['columns'] = ('distribution','account_number','tran_code','amount','effective_date','description',)
        tree.heading("#0", text='Tran ID', anchor=W)
        tree.column("#0",stretch=NO, width=10, anchor=W)
        tree.column('distribution', width=3, anchor=W)
        tree.heading('distribution', text='Distr')        
        tree.column('account_number', width=40, anchor=W)
        tree.heading('account_number', text='Account')        
        tree.column('tran_code', width=25, anchor=E)
        tree.heading('tran_code', text='Tran Code')       
        tree.column('amount', width=30, anchor=E)
        tree.heading('amount', text='Amount')       
        tree.column('effective_date', width=40, anchor=W)
        tree.heading('effective_date', text='Eff Dt')        
        tree.column('description', width=100, anchor=W)
        tree.heading('description', text='Description')
        ttk.Style().configure("Treeview", font=NORM_FONT, background='grey',
                              foreground='white',fieldbackground=BACKGROUND_COLOR)
        tree.grid(row=4, column=0, columnspan=6, ipadx=150, ipady=120, pady=20)
        # Class Object Array
        paymentArray = []

        ##########################################################
        # Method opens file explorer for user to select input file
        ##########################################################
        def selectFile(self):
            global input_file_name
            # Add adjustment to show all forms of excel files - xlsm, xls, etc
            filename = askopenfilename(filetypes = [("Excel Files","*.xlsx; *.xlsm")],
                               title = "Choose a file:")
            path_to_file = str(filename)

            if filename:
                mylist = filename.split("/")
                mylistSize = len(mylist)
                filename = mylist[mylistSize-1]

                i = 0
                filepath = ""
                while (i < mylistSize - 1):
                    filepath = filepath + mylist[i] + "/"
                    i = i + 1
                input_file_name = filename
                openFile(path_to_file)
            else:
                popupmsg("There was an error opening the file you\nselected or a file was not selected.")

        ########################################################
        # Method opens input file and appends paymentArray with
        # values from the input file
        ########################################################
        def openFile(path_to_file):       
            wb = openpyxl.load_workbook(path_to_file)
            sheet = wb.get_sheet_by_name('Payment Sheet') # Add functionality to select sheet
            end_of_sheet_range = get_column_letter(sheet.max_column) + str(sheet.max_row)
            
            # Fill array with excel data
            for rowOfCellObjects in sheet['A1':end_of_sheet_range]:
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) is datetime.datetime:
                        paymentArray.append(str(cellObj.value.strftime("%m/%d/%Y")))
                    elif type(cellObj.value) is type(None):
                        paymentArray.append("")
                    else:
                        paymentArray.append(cellObj.value)
            # Call method to fill table in frame with array values
            entryFill(self, paymentArray)

        ##########################################################
        # Method fills table in entry frame with the array values
        # and fills text window with debit / credit totals
        ##########################################################
        def entryFill(self, paymentArray):
            global debit_entry_total
            global credit_entry_total
            global input_transaction_count

            if len(paymentArray) % 6 == 0:
                input_transaction_count = len(paymentArray) / 6
                for i in range(1, (len(paymentArray)+1)):
                    if i%6==0:
                        tree.insert("","end",values = (str(paymentArray[i - 6]), str(paymentArray[i - 5]), str(paymentArray[i - 4]),\
                                                       str(paymentArray[i - 3]), str(paymentArray[i - 2]), str(paymentArray[i - 1])))
                        if paymentArray[i - 6] in DEBIT_TRAN_CODE:
                            debit_entry_total += paymentArray[i - 3]
                        elif paymentArray[i - 6] in CREDIT_TRAN_CODE:
                            credit_entry_total += paymentArray[i - 3]
                        else:
                            print('this isn\'t working how you think it is')

            else:   
                popupmsg('There was an issue with loading your transactions.\nPlease make sure you\'re input file has the correct\namount of entries and each part of the entry is\ninput correctly.')
            
            entry_data_label.configure(text=set_transaction_text(self))

        ########################################################
        # GUI method resets interface: empties the entries from 
        # table, clears array, and resets credit / debit totals.
        ########################################################
        def clear_payment_box(self):
            global debit_entry_total
            global credit_entry_total
            global input_file_name
            global input_transaction_count
            
            # Remove items from tree table in entry view frame
            x = tree.get_children()
            if x != '()':
                for child in x:
                    tree.delete(child)
            # Clear array
            del paymentArray[:]

            # Clear d / c totals & reset text
            debit_entry_total = 0.00
            credit_entry_total = 0.00
            input_file_name = "No File Selected"
            input_transaction_count = 0
            entry_data_label.configure(text=set_transaction_text(self))

        ###########################################################
        # Method receives a x, y coordinate variable, moves to the
        # coordinate and clicks it
        ###########################################################
        def select_position(coordinate_variable):
            pyautogui.moveTo(coordinate_variable, duration=.5)
            pyautogui.click()

        ################################################################
        # Method receives paymentArray to input values into Sparak.
        # Starts from Input Transactions screen then selects the green
        # 'New Transaction' button to bring up the 'New Transaction 
        # window. In New Tran window, primary for-loop iterates through
        # paymentArray and enters values or tabs based upon 6-digit
        # counter. 6-digit counter is reset when an entry is confirmed. 
        ################################################################
        def enter_into_sparak(self, paymentArray):
            if(len(paymentArray) == 0):
                popupmsg('No entries available to enter. Please\nload entries to enter into Sparak.')
            else:
                os.chdir(WRK_DIR)
                add_entry_button_location = pyautogui.locateOnScreen('plus_sign.png')
                if add_entry_button_location != None:
                    add_entry_button_location = pyautogui.center(add_entry_button_location)
                    counter = 1
                    select_position(add_entry_button_location)
                    pyautogui.PAUSE

                    
                    for item in paymentArray:
                        if counter < 6:
                            pyautogui.typewrite(str(item))      # Input array value into Trans Input
                            pyautogui.PAUSE = TIME_INTERVAL     # Forcing the time_inverval to a small number speeds up the process
                            pyautogui.press('tab')              # Move to next entry area
                        elif counter == 6:
                            pyautogui.typewrite(str(item))      # Input Description
                            pyautogui.PAUSE = TIME_INTERVAL     # Forcing the time_inverval to a small number speeds up the process
                            pyautogui.press('tab')              # Move to 'Repeat Distribution' box
                            pyautogui.PAUSE = TIME_INTERVAL
                            pyautogui.press('tab')              # Move to ok_button
                            pyautogui.PAUSE = TIME_INTERVAL
                            pyautogui.press('return')           # This pushes the entry
                            counter = 0                         # Reset counter for next entry
                        counter += 1     
                else:
                    popupmsg('I could not find Sparak\'s Transaction Input add entry\nbutton. Make sure Sparak is visible.')
        
        #####################################################
        # Method deletes a user input number of transactions
        #####################################################
        def delete_sparak_entries(self):
            os.chdir(WRK_DIR)
            delete_entry_button_location = pyautogui.locateOnScreen('delete_sign.png')
            if delete_entry_button_location != None:
                delete_entry_button_location = pyautogui.center(delete_entry_button_location)
                entryAmt = askinteger('Delete Transactions','Enter the number of transactions to delete from Sparak')
                if entryAmt is not None:
                    i = 0
                    while i < entryAmt:
                        pyautogui.click(delete_entry_button_location)
                        pyautogui.pause = TIME_INTERVAL
                        pyautogui.press('tab')
                        pyautogui.pause = TIME_INTERVAL
                        pyautogui.press('return')
                        i += 1
            else: 
                popupmsg('I could not find Sparak\'s Transaction Input screen.\nConfirm Sparak is visible and there are entries to delete.')

app = Sparak_Sloth()
app.geometry("640x640")     # Sets app window size
app.resizable(0,0)          # Keeps window size fixed
app.mainloop()
